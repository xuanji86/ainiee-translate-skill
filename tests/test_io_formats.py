import openpyxl
from ainiee_translate import io_dispatch, cache_io


def _roundtrip(tmp_path, fname, text):
    f = tmp_path / fname
    f.write_text(text, encoding="utf-8")
    proj = io_dispatch.parse_input(str(f))
    assert proj.count_items() >= 1
    for it in cache_io.iter_items(proj):
        cache_io.set_translation(proj, it.text_index, "T_" + (it.source_text or ""))
    out = tmp_path / "out"
    io_dispatch.export_project(proj, str(out), str(f))
    return proj, out


def test_srt_roundtrip(tmp_path):
    proj, out = _roundtrip(
        tmp_path, "a.srt",
        "1\n00:00:01,000 --> 00:00:02,000\nHello\n\n2\n00:00:03,000 --> 00:00:04,000\nWorld\n")
    assert proj.project_type == "Srt"
    assert "T_Hello" in (out / "a_translated.srt").read_text(encoding="utf-8")


def test_csv_roundtrip(tmp_path):
    proj, out = _roundtrip(tmp_path, "a.csv", "h1,h2\nHello,World\n")
    assert proj.project_type == "Csv"
    assert "T_Hello" in (out / "a_translated.csv").read_text(encoding="utf-8")


def test_po_roundtrip(tmp_path):
    proj, out = _roundtrip(tmp_path, "a.po", 'msgid "Hello"\nmsgstr ""\n')
    assert proj.project_type == "Po"
    assert "T_Hello" in (out / "a_translated.po").read_text(encoding="utf-8")


def test_xlsx_roundtrip_and_ambiguous_detection(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["A2"] = "Hello", "World"
    wb.save(tmp_path / "a.xlsx")
    proj = io_dispatch.parse_input(str(tmp_path / "a.xlsx"))
    assert proj.project_type == "Xlsx"          # .xlsx disambiguated to generic Xlsx
    assert proj.count_items() >= 1
    for it in cache_io.iter_items(proj):
        cache_io.set_translation(proj, it.text_index, "T_" + (it.source_text or ""))
    io_dispatch.export_project(proj, str(tmp_path / "out"), str(tmp_path / "a.xlsx"))
    assert (tmp_path / "out" / "a_translated.xlsx").is_file()


def test_registry_complete():
    assert len(io_dispatch.FORMATS) == 20                       # all light formats
    assert all(r and w for r, w in io_dispatch.FORMATS.values())
    assert {"epub", "txt", "md", "srt", "csv", "po", "xlsx", "json", "docx", "pptx"} \
        <= set(io_dispatch.supported_extensions())
